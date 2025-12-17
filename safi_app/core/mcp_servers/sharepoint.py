import os
import json
import logging
import requests
from datetime import datetime
from flask import session, current_app
from ...persistence.database import get_oauth_token, upsert_oauth_token

logger = logging.getLogger("sharepoint_mcp")

def get_valid_token():
    user_id = session.get('user_id')
    if not user_id:
        raise Exception("User not logged in.")
        
    token_data = get_oauth_token(user_id, 'microsoft')
    if not token_data:
        raise Exception("Microsoft account not connected. Please connect in Settings.")
        
    # Check expiration
    expires_at = token_data.get('expires_at')
    # If expires_at is None, we assume it's valid or let it fail downstream, but here we check if it IS set and past.
    if expires_at and expires_at < datetime.now():
        logger.info("Token expired, refreshing...")
        # Refresh logic
        url = "https://login.microsoftonline.com/common/oauth2/v2.0/token"
        data = {
            "client_id": current_app.config['MICROSOFT_CLIENT_ID'],
            "client_secret": current_app.config['MICROSOFT_CLIENT_SECRET'],
            "refresh_token": token_data['refresh_token'],
            "grant_type": "refresh_token"
        }
        r = requests.post(url, data=data)
        r.raise_for_status()
        new_tokens = r.json()
        
        access_token = new_tokens['access_token']
        # Some refreshes return a new refresh token, some don't
        refresh_token = new_tokens.get('refresh_token', token_data['refresh_token'])
        
        expires_in = new_tokens.get('expires_in', 3600)
        from datetime import timedelta
        new_expires_at = datetime.now() + timedelta(seconds=expires_in)

        upsert_oauth_token(
            user_id, 'microsoft', access_token, refresh_token, 
            new_expires_at,
            new_tokens.get('scope', token_data['scope'])
        )
        return access_token
        
    return token_data['access_token']

async def search_drive(query: str) -> str:
    try:
        token = get_valid_token()
        headers = {"Authorization": f"Bearer {token}"}
        
        # Graph API Search
        # Handle wildcard or empty search by listing root children instead
        if not query or query.strip() == "*":
             url = "https://graph.microsoft.com/v1.0/me/drive/root/children"
        else:
             url = f"https://graph.microsoft.com/v1.0/me/drive/root/search(q='{query}')"
             
        r = requests.get(url, headers=headers)
        r.raise_for_status()
        
        data = r.json()
        items = []
        for item in data.get('value', [])[:20]:
            items.append({
                "name": item.get('name'),
                "id": item.get('id'),
                "webUrl": item.get('webUrl'),
                "folder": "folder" in item,
                "size": item.get('size'),
                "lastModified": item.get('lastModifiedDateTime')
            })
            
        return json.dumps(items)
    except Exception as e:
        logger.error(f"SharePoint Search Error: {e}")
        return json.dumps({"error": str(e)})

async def list_folders(folder_path: str = "root") -> str:
    """Lists folders and files in a specific directory."""
    try:
        token = get_valid_token()
        headers = {"Authorization": f"Bearer {token}"}
        
        if folder_path == "root" or folder_path == "/" or not folder_path:
            url = "https://graph.microsoft.com/v1.0/me/drive/root/children"
        else:
            # Handle path based addressing (needs drive/root:/path:/children)
            clean_path = folder_path.strip("/")
            url = f"https://graph.microsoft.com/v1.0/me/drive/root:/{clean_path}:/children"

        r = requests.get(url, headers=headers)
        r.raise_for_status()
        
        data = r.json()
        items = []
        for item in data.get('value', []):
            items.append({
                "name": item.get('name'),
                "id": item.get('id'),
                "type": "folder" if "folder" in item else "file",
                "size": item.get('size'),
                "lastModified": item.get('lastModifiedDateTime')
            })
            
        return json.dumps(items)
    except Exception as e:
        logger.error(f"SharePoint List Folders Error: {e}")
        return json.dumps({"error": str(e)})

async def get_tree(max_depth: int = 2) -> str:
    """Returns a simplified tree structure of folders only."""
    # Note: Recursive traversal can be slow. We limit depth.
    # For a PoC, let's just do root + 1 level of folders
    try:
        token = get_valid_token()
        headers = {"Authorization": f"Bearer {token}"}
        
        async def fetch_children(url):
            try:
                r = requests.get(url, headers=headers)
                if r.status_code != 200: return []
                return r.json().get('value', [])
            except: return []

        root_items = await asyncio.to_thread(lambda: requests.get("https://graph.microsoft.com/v1.0/me/drive/root/children", headers=headers).json().get('value', []))
        
        tree = []
        for item in root_items:
            if "folder" in item:
                node = {"name": item["name"], "type": "folder", "children": []}
                # Level 1 fetch
                # If max_depth > 1...
                # sub_url = f"https://graph.microsoft.com/v1.0/me/drive/items/{item['id']}/children"
                # sub_items = ...
                # node["children"] = sub_items... 
                tree.append(node)
            
        return json.dumps(tree) # Simplified for speed
    except Exception as e:
        return json.dumps({"error": str(e)})


async def read_item(item_id: str) -> str:
    try:
        token = get_valid_token()
        headers = {"Authorization": f"Bearer {token}"}
        
        # Get metadata first to check name/extension
        meta_url = f"https://graph.microsoft.com/v1.0/me/drive/items/{item_id}"
        meta_r = requests.get(meta_url, headers=headers)
        meta_r.raise_for_status()
        metadata = meta_r.json()
        name = metadata.get('name', '').lower()
        
        # Get content stream
        url = f"https://graph.microsoft.com/v1.0/me/drive/items/{item_id}/content"
        r = requests.get(url, headers=headers)
        r.raise_for_status()
        content_bytes = r.content

        # --- CONTENT PROCESSING ---
        
        if name.endswith('.xlsx'):
            import openpyxl
            from io import BytesIO
            wb = openpyxl.load_workbook(BytesIO(content_bytes), data_only=True)
            text_output = []
            for sheet in wb.sheetnames:
                text_output.append(f"--- Sheet: {sheet} ---")
                ws = wb[sheet]
                # Limit rows for token safety
                for i, row in enumerate(ws.iter_rows(values_only=True)):
                    if i > 50: break 
                    text_output.append(str(row))
            return "\n".join(text_output)

        elif name.endswith('.docx'):
            import docx
            from io import BytesIO
            doc = docx.Document(BytesIO(content_bytes))
            return "\n".join([p.text for p in doc.paragraphs])
            
        elif name.endswith('.pdf'):
            import pypdf
            from io import BytesIO
            pdf = pypdf.PdfReader(BytesIO(content_bytes))
            text = []
            for page in pdf.pages:
                text.append(page.extract_text())
            return "\n".join(text)

        # Fallback to Text
        try:
             return content_bytes.decode('utf-8')
        except:
             return f"[Binary Content: {len(content_bytes)} bytes - File type not natively supported for text extraction]"
             
    except Exception as e:
        logger.error(f"SharePoint Read Error: {e}")
        return json.dumps({"error": str(e)})

async def upload_item(name: str, content: str) -> str:
    try:
        token = get_valid_token()
        headers = {
            "Authorization": f"Bearer {token}",
            "Content-Type": "text/plain"
        }
        
        # Upload to root
        url = f"https://graph.microsoft.com/v1.0/me/drive/root:/{name}:/content"
        r = requests.put(url, headers=headers, data=content.encode('utf-8'))
        r.raise_for_status()
        
        return json.dumps(r.json())
    except Exception as e:
        logger.error(f"SharePoint Upload Error: {e}")
        return json.dumps({"error": str(e)})

async def search_sites(query: str) -> str:
    try:
        token = get_valid_token()
        headers = {"Authorization": f"Bearer {token}"}
        
        # Graph API Site Search
        # Note: use "v1.0/sites?search={query}"
        url = f"https://graph.microsoft.com/v1.0/sites?search={query}"
        r = requests.get(url, headers=headers)
        r.raise_for_status()
        
        data = r.json()
        items = []
        for site in data.get('value', [])[:10]:
            items.append({
                "name": site.get('displayName'),
                "id": site.get('id'),
                "webUrl": site.get('webUrl'),
                "description": site.get('description', '')
            })
            
        return json.dumps(items)
    except Exception as e:
        logger.error(f"SharePoint Site Search Error: {e}")
        return json.dumps({"error": str(e)})

async def search_site_drive(site_id: str, query: str) -> str:
    try:
        token = get_valid_token()
        headers = {"Authorization": f"Bearer {token}"}
        
        # 1. Get the default Drive ID for the site (often simpler than finding all drives)
        # However, graph supports /sites/{id}/drive/root/search directly usually.
        # Let's try direct search first.
        url = f"https://graph.microsoft.com/v1.0/sites/{site_id}/drive/root/search(q='{query}')"
        
        r = requests.get(url, headers=headers)
        r.raise_for_status()
        
        data = r.json()
        items = []
        # Reuse mapping logic if possible or just dump relevant fields
        for item in data.get('value', [])[:10]:
            items.append({
                "name": item.get('name'),
                "id": item.get('id'),
                "webUrl": item.get('webUrl'),
                "folder": "folder" in item,
                "lastModified": item.get('lastModifiedDateTime')
            })
            
        return json.dumps(items)
    except Exception as e:
        logger.error(f"SharePoint Site Drive Search Error: {e}")
        return json.dumps({"error": str(e)})

